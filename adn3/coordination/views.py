from django.http import HttpResponse, Http404
from django.shortcuts import render, get_object_or_404
from django.views import generic
from django.views import View
import tempfile
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from adn3.constants import YEAR, SEMESTER

from courses.models import Course
from misc.models import Software


class CoordinationIndexView(View):
    def get(self, request):
        current_courses = Course.objects.filter(semester=SEMESTER, year=YEAR)
        return render(request, 'coordination/coordination_index.html', {
            'current_courses': current_courses,
            'software_list': Software.objects.all(),
        })


class PreRegistrationExcelView(View):
    def get(self, request):
        course = self.get_course()
        software = self.get_software()

        pr_list = course.preregistration_set.filter(software=software).all()

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['E2'] = str(course)
        ws['E3'] = 'Preinscripciones software {software} para {date}'.format(
            software=software.name,
            date='20170311_12:39:17'
        )

        ws['B5'] = 'Nro'
        ws['D5'] = 'ROL USM'
        ws['F5'] = 'A. Paterno'
        ws['G5'] = 'Nombres'
        ws['I5'] = 'PA'
        ws['K5'] = 'P1'
        ws['M5'] = 'P2'
        ws['O5'] = 'P3'
        ws['Q5'] = 'P4'
        ws['S5'] = 'P5'
        ws['U5'] = 'Exp'
        ws['W5'] = 'Par'

        for i, pr in enumerate(pr_list):
            ws['B%s' % (6 + i)] = i
            ws['C%s' % (6 + i)] = '{{"'
            ws['D%s' % (6 + i)] = pr.student.rol
            ws['E%s' % (6 + i)] = '","'
            ws['F%s' % (6 + i)] = pr.student.user.last_name.split()[0]
            ws['G%s' % (6 + i)] = pr.student.user.first_name
            ws['H%s' % (6 + i)] = '",'
            ws['I%s' % (6 + i)] = pr.student.usm_priority
            ws['J%s' % (6 + i)] = ','
            ws['K%s' % (6 + i)] = pr.first_preference
            ws['L%s' % (6 + i)] = ','
            ws['M%s' % (6 + i)] = pr.second_preference
            ws['N%s' % (6 + i)] = ','
            ws['O%s' % (6 + i)] = pr.third_preference
            ws['P%s' % (6 + i)] = ','
            ws['Q%s' % (6 + i)] = pr.fourth_preference
            ws['R%s' % (6 + i)] = ','
            ws['S%s' % (6 + i)] = pr.fifth_preference
            ws['T%s' % (6 + i)] = ','
            ws['U%s' % (6 + i)] = 0
            ws['V%s' % (6 + i)] = '},'
            ws['W%s' % (6 + i)] = 0

        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=preinscripciones.xlsx'
        return response

    def get_course(self):
        course_pk = self.request.GET.get('course')

        if course_pk is None:
            raise Http404()

        return get_object_or_404(Course, pk=course_pk)

    def get_software(self):
        software_pk = self.request.GET.get('software')

        if software_pk is None:
            raise Http404()

        return get_object_or_404(Software, pk=software_pk)
