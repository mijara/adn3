import openpyxl
from openpyxl.writer.excel import save_virtual_workbook


class GradeMatrix(object):
    def __init__(self):
        self.student_names = []
        self.tests = {}
        self.tests_ordering = []

    def set_user(self, name):
        self.student_names.append(name)

    def add(self, test_name, grade):
        if test_name not in self.tests:
            self.tests[test_name] = []
            self.tests_ordering.append(test_name)
        self.tests[test_name].append(grade)

    def _row(self, index, student):
        row = [student]
        for test in self.tests_ordering:
            row.append(self.tests[test][index] or 0)
        return row

    def iter_rows(self):
        for i, student in enumerate(self.student_names):
            yield self._row(i, student)

    def render(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Nombre'] + self.tests_ordering)

        for row in self.iter_rows():
            ws.append(row)

        return save_virtual_workbook(wb)
