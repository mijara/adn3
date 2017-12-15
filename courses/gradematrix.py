import openpyxl
from openpyxl.writer.excel import save_virtual_workbook


class GradeMatrix(object):
    def __init__(self):
        self.student_names = []
        self.tests = {}
        self.pretests = {}
        self.tests_ordering = []
        self.pretests_ordering = []

    def set_user(self, name):
        self.student_names.append(name)

    def add(self, test_name, grade):
        if test_name not in self.tests:
            self.tests[test_name] = []
            self.tests_ordering.append(test_name)
        self.tests[test_name].append(grade)

    def add_pretest(self, name, grade):
        if name not in self.pretests:
            self.pretests[name] = []
            self.pretests_ordering.append(name)
        self.pretests[name].append(grade)

    def _row(self, index, student):
        row = [student]
        for test in self.tests_ordering:
            row.append(self.tests[test][index] or 0)
        for pretest in self.pretests_ordering:
            row.append(self.pretests[pretest][index] or 0)
        return row

    def iter_rows(self):
        for i, student in enumerate(self.student_names):
            yield self._row(i, student)

    def render(self):
        wb = openpyxl.Workbook()

        self._render_resume(wb.active)

        return save_virtual_workbook(wb)

    def _render_resume(self, ws):
        ws.title = "Resumen"
        ws.append(['Nombre'] + self.tests_ordering + self.pretests_ordering)

        for row in self.iter_rows():
            ws.append(row)
