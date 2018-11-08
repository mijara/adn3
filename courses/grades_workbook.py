import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

from attendance.models import Attendance
from courses.table import Table
from pretests.models import PretestUpload
from tests.models import Answer, StudentsAnswers, Version


def clean_invalid_characters(string):
    invalid_characters = ['*', ':', '/', '\\', '?', '[', ']']
    for character in invalid_characters:
        string = string.replace(character, ' ')

    return string


def _questions_row(student_answers, version):
    row = []
    for question in version.question_set.all():
        answer = Answer.objects.filter(student=student_answers.student, question=question)

        if answer.exists():
            answer = answer.first()
            row.append(question.score if answer.correct else 0)

    # add final grade of this test to the row.
    row.append(student_answers.qualification)
    return row


def _append_test_version(student, personal, test, wb, sheets):
    sa = StudentsAnswers.objects.filter(student=student.user, version__test=test)
    #  print(sa)

    if sa.exists():
        sa = sa.first()

        if sa.version.pk not in sheets:
            title = clean_invalid_characters(test.name + ' - ' + sa.version.get_letter())
            sheets[sa.version.pk] = wb.create_sheet(title)
            header = ['Rol', 'Nombres', 'Apellidos']
            header += ['P%d' % (i + 1) for i in range(sa.version.question_set.count())] + ['Nota']
            sheets[sa.version.pk].append(header)

        sheets[sa.version.pk].append(personal + _questions_row(sa, sa.version))

        return sa.qualification
    else:
        return 0


def _append_tests(course, wb, resume):
    sheets = {}

    for student in course.get_students():
        rol = student.rol
        personal = [rol, student.user.first_name, student.user.last_name]

        # add test qualification to general view.
        if rol not in resume:
            resume[rol] = personal[:]

        for test in course.test_set.all():
            grade = _append_test_version(student, personal, test, wb, sheets)
            resume[rol].append(grade)


def _append_pretests(course, resume):
    for pretest in course.pretest_set.all():
        for student in course.get_students():
            upload = PretestUpload.objects.filter(pretest=pretest, student=student)

            if upload.exists():
                upload = upload.first()
                resume[student.rol].append(upload.qualification)
            else:
                resume[student.rol].append(0)


def _append_assistance(course, resume):
    sessions_count = course.session_set.count()

    for agenda in course.agenda_set.all():
        for student in agenda.inscriptions.all():
            attendance = Attendance.objects.filter(agenda=agenda, user=student). \
                exclude(attended=Attendance.NO_ATTENDED).count()
            resume[student.student.rol].append(attendance * 100 / sessions_count)


def generate_course_grades(course):
    wb = openpyxl.Workbook()

    resume_ws = wb.active
    header = ['Rol', 'Nombres', 'Apellidos']
    header += ['C%d' % (i + 1) for i in range(course.test_set.count())]
    header += ['P%d' % (i + 1) for i in range(course.pretest_set.count())]
    header += ['A', 'F']
    resume_ws.append(header)

    resume_by_student = {}
    _append_tests(course, wb, resume_by_student)
    _append_pretests(course, resume_by_student)
    _append_assistance(course, resume_by_student)

    # get the grade configuration.
    gc = course.grades_config

    # add the resume table to the worksheet.
    for resume in resume_by_student.values():
        resume = [0 if v is None else v for v in resume]

        len_tests = course.test_set.count()
        len_pretests = course.pretest_set.count()

        tests = sum(resume[3:(3 + len_tests)]) / len_tests

        if len_pretests != 0:
            pretests = sum(resume[(3 + len_tests):(3 + len_tests + len_pretests)]) / len_pretests
        else:
            pretests = 0

        assistance = resume[-1]

        final = tests * gc.grade_tests / 100 + pretests * gc.grade_pretests / 100 + \
                assistance * gc.grade_assistance / 100
        resume_ws.append(resume + [final])

    return save_virtual_workbook(wb)


def generate_course_grades_v2(course):
    resume_table = Table()

    sessions_total = course.session_set.count()

    tests = [t for t in course.test_set.all()]
    pretests = [p for p in course.pretest_set.all()]

    resume_table.fill_columns(["Nombres", "Apellidos", "Rol"])

    resume_table.fill_columns([f'C{i+1}' for i in range(len(tests))])
    resume_table.fill_columns([f'P{i+1}' for i in range(len(pretests))])

    for student in course.get_students():
        # personal data
        resume_table.set(row=student.user.email, column=f'Nombres', value=student.user.first_name)
        resume_table.set(row=student.user.email, column=f'Apellidos', value=student.user.last_name)
        resume_table.set(row=student.user.email, column=f'Rol', value=student.rol)

        # tests.
        student_answers = StudentsAnswers.objects.filter(student=student.user, version__test__course=course).all()
        for student_answer in student_answers:
            resume_table.set(
                row=student.user.email,
                column=f'C{tests.index(student_answer.version.test) + 1}',
                value=student_answer.qualification)

        # pretests.
        pretest_uploads = PretestUpload.objects.filter(student=student, pretest__course=course).all()
        for upload in pretest_uploads:
            resume_table.set(
                row=student.user.email,
                column=f'P{pretests.index(upload.pretest) + 1}',
                value=upload.qualification)

        # assistance.
        attendance = Attendance.objects.filter(user=student.user, agenda__course=course). \
            exclude(attended=Attendance.NO_ATTENDED).count()

        resume_table.set(
            row=student.user.email,
            column="Asistencia",
            value=int(attendance * 100 / sessions_total))

    resume_table.fill_blanks()

    # generate excel.
    wb = openpyxl.Workbook()
    resume_ws = wb.active
    resume_ws.title = "General"

    resume_table.as_excel(resume_ws)

    for test in tests:
        for version in test.version_set.all():
            title = clean_invalid_characters(version.test.name + ' - ' + version.get_letter())
            ws = wb.create_sheet(title)
            generate_version_excel(course, version).as_excel(ws)

    ws = wb.create_sheet('Asistencia')
    generate_assistance_excel(course).as_excel(ws)

    return save_virtual_workbook(wb)


def generate_version_excel(course, version):
    table = Table()

    questions = [q for q in version.question_set.all()]

    for student in course.get_students():
        sa = StudentsAnswers.objects.filter(student=student.user, version=version)
        if not sa.exists():
            table.add_row(student.user.email)
            continue
        sa = sa.get()

        answers = Answer.objects.filter(student=student.user, question__version=version).all()

        for answer in answers:
            table.set(row=student.user.email, column=f'P{questions.index(answer.question) + 1}', value=answer.correct)

        table.set(row=student.user.email, column=f'Nota', value=sa.qualification)

    table.fill_columns([f'P{questions.index(q) + 1}' for q in questions] + ['Nota'])
    table.fill_blanks()

    return table


def generate_assistance_excel(course):
    table = Table()

    sessions = [s for s in course.session_set.all()]
    table.fill_columns([s.number for s in sessions])

    for student in course.get_students():
        for session in sessions:
            attendance = Attendance.objects.filter(session=session, user=student.user)

            if not attendance.exists():
                continue

            attendance = attendance.first()

            table.set(row=student.user.email, column=session.number, value=attendance.get_attended_display())

    table.fill_blanks()

    return table
