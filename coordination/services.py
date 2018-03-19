from typing import List

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from datetime import datetime

from adn3.services import get_period_year, get_period_semester
from courses.models import Course
from polls.models import Poll


def dayblock_to_adn2_format(day):
    """
    Transform the internal day-block representation to ADN2's representation:
    {block}{day}, both starting at 1.
    """
    if day is None:
        return '11'
    day, block = day.split("-")
    return "{}{}".format(int(block) + 1, int(day) + 1)


def generate_excel(course, software, pr_list):
    wb = openpyxl.Workbook()
    ws = wb.active

    dt = datetime.now()

    ws['E2'] = str(course)
    ws['E3'] = 'Preinscripciones software {software} para {date}'.format(
        software=software.name,
        # format example: '20170311_12:39:17'
        date=dt.strftime("%Y%m%d_%H:%M:%S")
    )

    ws['B5'] = 'Nro'
    ws['D5'] = 'ROL USM'
    ws['F5'] = 'A. Paterno'
    ws['G5'] = 'Nombres'
    ws['I5'] = 'PA'
    ws['K5'] = 'P1'
    ws['M5'] = 'P2'
    ws['O5'] = 'P3'
    ws['Q5'] = 'P4'
    ws['S5'] = 'P5'
    ws['U5'] = 'Exp'
    ws['W5'] = 'Par'

    for i, pr in enumerate(pr_list):
        ws['B%s' % (6 + i)] = i + 1
        ws['C%s' % (6 + i)] = '{{"' if i == 0 else '{"'
        # ws['D%s' % (6 + i)] = pr.student.rol
        ws['D%s' % (6 + i)] = pr.rol
        ws['E%s' % (6 + i)] = '","'
        # ws['F%s' % (6 + i)] = pr.student.user.last_name.split()[0]
        ws['F%s' % (6 + i)] = pr.last_names.split()[0]
        # ws['G%s' % (6 + i)] = pr.student.user.first_name
        ws['G%s' % (6 + i)] = pr.first_name
        ws['H%s' % (6 + i)] = '",'
        # ws['I%s' % (6 + i)] = pr.student.usm_priority
        ws['I%s' % (6 + i)] = pr.usm_priority
        ws['J%s' % (6 + i)] = ','
        ws['K%s' % (6 + i)] = dayblock_to_adn2_format(pr.first_preference)
        ws['L%s' % (6 + i)] = ','
        ws['M%s' % (6 + i)] = dayblock_to_adn2_format(pr.second_preference)
        ws['N%s' % (6 + i)] = ','
        ws['O%s' % (6 + i)] = dayblock_to_adn2_format(pr.third_preference)
        ws['P%s' % (6 + i)] = ','
        ws['Q%s' % (6 + i)] = dayblock_to_adn2_format(pr.fourth_preference)
        ws['R%s' % (6 + i)] = ','
        ws['S%s' % (6 + i)] = dayblock_to_adn2_format(pr.fifth_preference)
        ws['T%s' % (6 + i)] = ','
        ws['U%s' % (6 + i)] = int(pr.previous_experience)
        ws['V%s' % (6 + i)] = '},' if i + 1 < len(pr_list) else '}}'
        ws['W%s' % (6 + i)] = pr.parallel

    return save_virtual_workbook(wb)


def generate_polls_excel(course, poll_list):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['E2'] = 'Encuesta fin de semestre'
    ws['E3'] = str(course)

    ws['B5'] = 'Nro'
    ws['C5'] = 'Año'
    ws['D5'] = 'Semestre'

    letters = "EFGHIJKLMNOPQRSTUVWXYZ"
    for i, q in enumerate(Poll._meta.get_fields()[3:]):
        ws['%s5' % letters[i]] = q.verbose_name

    for i, pll in enumerate(poll_list):
        ws['B%s' % (6 + i)] = i
        ws['C%s' % (6 + i)] = pll.course.year
        ws['D%s' % (6 + i)] = pll.course.semester

        p_dict = pll.__dict__
        keys = list(p_dict.keys())[4:-1]

        for j, q in enumerate(keys):
            ws['%s%s' % (letters[j], (6 + i))] = p_dict[q]

    return save_virtual_workbook(wb)


def generate_preregistrations_overview_excel(courses: List[Course]):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append([
        'Curso',
        'Sede',
        'Código',
        'Año-Semestre',
        'Estudiantes Preinscritos'
    ])

    for course in courses:
        ws.append([
            course.name,
            course.campus.name,
            course.code,
            "%s-%s" % (get_period_year(), get_period_semester()),
            course.flyin_set.count()
        ])

    return save_virtual_workbook(wb)
